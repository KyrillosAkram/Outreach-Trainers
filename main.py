print("\n\rLoading... 0/4",end='')
from bs4 import BeautifulSoup
print("\rLoading... 1/4",end='')
import requests
print("\rLoading... 2/4",end='')
import json
print("\rLoading... 3/4",end='')
import openpyxl

header_request={
'Host': 'www.osha.gov',
'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:75.0) Gecko/20100101 Firefox/75.0',
'Accept': '*/*',
'Accept-Language':'en-US,en;q=0.5',
'Accept-Encoding': 'gzip, deflate, br',
'Connection': 'keep-alive',
'Referer': 'https://www.osha.gov/dte/outreach/outreach_trainers_webworker.js',
'Pragma': 'no-cache',
'Cache-Control': 'no-cache',
'TE': 'Trailers'
}
xlheader=['FIRST NAME', 'LAST NAME', 'EMAIL' ,'PHONE']
print('\rget request send ...')
resposne=requests.get('https://www.osha.gov/dte/outreach/outreach_trainers.json',headers=header_request)
print('\rstart data extraction ...')
resposne=BeautifulSoup(resposne.text,'lxml')
print('\rreforming data')
trainers=json.loads(resposne.html.body.p.get_text())

ws=openpyxl.Workbook()
sheet=ws.get_sheet_by_name("Sheet")
#sheet=ws.create_sheet()
sheet.cell(row=1,column=1).value=xlheader[0]
sheet.cell(row=1,column=2).value=xlheader[1]
sheet.cell(row=1,column=3).value=xlheader[2]
sheet.cell(row=1,column=4).value=xlheader[3]

rownum=2

for trainer in trainers:
    print("\rrow %0.5d formed"%(rownum),end='')
    #   no
    sheet.cell(row=rownum,column=1).value=trainer['FIRST NAME']
    #   name english
    sheet.cell(row=rownum,column=2).value=trainer['LAST NAME']
    #   name chinese
    sheet.cell(row=rownum,column=3).value=trainer['EMAIL']
    #   link
    sheet.cell(row=rownum,column=4).value=trainer['PHONE']
    #   increament rownum
    rownum=rownum+1

ws.save("trainers.xlsx")

print("\n\rdata is saved in trainers.xlsx\n\a")